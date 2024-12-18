import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from tempfile import NamedTemporaryFile

# Define column mapping
COLUMN_MAPPING = {
    "Sl. No.": "Sr #",
    "Date of Notice": "",
    "Time of Notice": "",
    "Name and Address of Injured Person": "Name",
    "Sex": "Gender",
    "Age": "Age",
    "Insurance No.": "",
    "Shift, department and Occupation of the employee ": "",
    "details of injury_cause": "Age",
    "nature": "Age",
    "date": "",
    "time": "Age",
    "What exactly was the injured person doing at the time of accident": "",
    "Name,occupation,address and signature or the thumb impression of the person(s) giving notice": "",
    "Signature and designation of the person who makes the entry in the Accident Book": "",
    "Name, address and occupation of two witnesses ": "",
    "Remarks, if any": ""
}

def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""

def combine_headers(row_11, row_12):
    """Combine headers from two rows to handle nested headers."""
    combined_headers = []
    for col_idx, header_row_11 in enumerate(row_11):
        header_row_12 = row_12[col_idx] if col_idx < len(row_12) else None
        if header_row_11 and header_row_12:
            combined_headers.append(f"{header_row_11.strip()}_{header_row_12.strip()}")
        elif header_row_11:
            combined_headers.append(header_row_11.strip())
        elif header_row_12:
            combined_headers.append(header_row_12.strip())
        else:
            combined_headers.append("")
    return combined_headers
def accident_process_excel(accident_file_path, master_file_path, output_file_path):
    # Load the input workbook (accident.xlsx)
    wb_accident = load_workbook(accident_file_path)
    sheet_accident = wb_accident.active  # Assuming first sheet in accident.xlsx

    # Load the master workbook (master.xlsx)
    wb_master = load_workbook(master_file_path)
    sheet_master = wb_master.active  # Assuming first sheet in master.xlsx

    # Combine headers from row 11 and 12 in the accident file
    row_11 = [cell.value for cell in sheet_accident[11]]
    row_12 = [cell.value for cell in sheet_accident[12]]
    combined_headers = combine_headers(row_11, row_12)

    # Normalize combined headers
    accident_headers = {normalize_string(header): idx + 1 for idx, header in enumerate(combined_headers)}

    # Get headers from row 3 in master.xlsx and normalize
    master_headers = {normalize_string(cell.value): idx for idx, cell in enumerate(sheet_master[3], start=1)}

    # Unmerge all merged cells in accident.xlsx before clearing
    merged_cells = list(sheet_accident.merged_cells)  # Store merged cell ranges
    for merged_range in merged_cells:
        sheet_accident.unmerge_cells(str(merged_range))

    # Clear content in accident.xlsx from row 13 onwards
    for row in sheet_accident.iter_rows(min_row=13):
        for cell in row:
            cell.value = None

    # Insert data from master.xlsx into accident.xlsx starting from row 13
    row_idx = 13
    data_inserted = False  # Flag to check if any data is inserted
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):  # Data starts from row 4
        accident_row_data = {}  # Temporary storage for the accident row data
        for accident_col, master_col in COLUMN_MAPPING.items():
            if master_col:  # Skip mappings with empty master columns
                accident_col_normalized = normalize_string(accident_col)
                master_col_normalized = normalize_string(master_col)
                if accident_col_normalized in accident_headers and master_col_normalized in master_headers:
                    accident_col_idx = accident_headers[accident_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    if accident_col == "Insurance No." and not master_row[master_col_idx - 1]:
                        continue  # Skip this column if "Insurance No." has no value
                    accident_row_data[accident_col_idx] = master_row[master_col_idx - 1]

        # Only write the row if any data exists
        if accident_row_data:
            for col_idx, value in accident_row_data.items():
                cell = sheet_accident.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align data
            row_idx += 1
            # data_inserted = True

    # If no data was inserted, write "NO ANY ACCIDENT FOR THE MONTH OF OCT 2024" in row 18, column D
    if not data_inserted:
        for row in sheet_accident.iter_rows(min_row=13):
            for cell in row:
                cell.value = None
        message_cell = sheet_accident.cell(row=18, column=4)
        message_cell.value = "NO ANY ACCIDENT FOR THE MONTH OF NOV 2024"
        message_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Apply borders to all cells under row 12 for columns with headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in accident_headers.values():
        for row_idx in range(13, sheet_accident.max_row + 1):
            cell = sheet_accident.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge cells only up to row 12
    for merged_range in merged_cells:
        start_row = merged_range.min_row
        if start_row <= 12:
            sheet_accident.merge_cells(str(merged_range))

    # Save the updated accident.xlsx file
    wb_accident.save(output_file_path)
    print(f"Data inserted and borders applied successfully. New file saved as: {output_file_path}")

    return output_file_path
