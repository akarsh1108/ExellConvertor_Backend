import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# Define column mapping
COLUMN_MAPPING = {
    "Sl.No": "Sr #",
    "Name of  workman": "Name",
    "Father's/ Husband's  Name": "Father Name",
    "Designation / nature of employment": "Designation",
    "Act/Ommision for which fine imposed": "",
    "Date od Offence": "",
    "Name of person in whose presence employee's explanation was heared": "",
    "Wage period and wage payable": "",
    "Amount of fine imposed": "",
    "Date on which fine realised": "",
    "Remarks": "",
}

def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""
def combine_headers(row_11, row_12):
    """Combine headers from two rows to handle nested headers."""
    combined_headers = []
    for col_idx, header_row_11 in enumerate(row_11):
        header_row_12 = row_12[col_idx] if col_idx < len(row_12) else None
        if header_row_11 and header_row_12:
            combined_headers.append(f"{header_row_11.strip()}_{header_row_12.strip()}")
        elif header_row_11:
            combined_headers.append(header_row_11.strip())
        elif header_row_12:
            combined_headers.append(header_row_12.strip())
        else:
            combined_headers.append("")
    return combined_headers
def fine_process_excel(base_file_path, temp_master_file_path, temp_output_file):
    """Process the Fine Excel file with data from the master file."""
    # Load the input workbook (Fine.xlsx)
    wb_fine = load_workbook(base_file_path)
    sheet_fine = wb_fine.active  # Assuming the first sheet in Fine.xlsx

    # Load the master workbook (master.xlsx)
    wb_master = load_workbook(temp_master_file_path)
    sheet_master = wb_master.active  # Assuming the first sheet in master.xlsx

    # Combine headers from row 11 and 12 in the Fine file
    row_11 = [cell.value for cell in sheet_fine[11]]
    row_12 = [cell.value for cell in sheet_fine[12]]
    combined_headers = combine_headers(row_11, row_12)

    # Normalize combined headers
    fine_headers = {normalize_string(header): idx + 1 for idx, header in enumerate(combined_headers)}

    # Get headers from row 3 in master.xlsx and normalize
    master_headers = {normalize_string(cell.value): idx for idx, cell in enumerate(sheet_master[3], start=1)}

    # Unmerge all merged cells in Fine.xlsx before clearing
    merged_cells = list(sheet_fine.merged_cells)  # Store merged cell ranges
    for merged_range in merged_cells:
        sheet_fine.unmerge_cells(str(merged_range))

    # Clear content in Fine.xlsx from row 14 onwards
    for row in sheet_fine.iter_rows(min_row=14):
        for cell in row:
            cell.value = None

    # Insert data from master.xlsx into Fine.xlsx starting from row 14
    row_idx = 14
    data_inserted = False  # Flag to check if any data is inserted
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):  # Data starts from row 4
        fine_row_data = {}  # Temporary storage for the fine row data
        skip_row = False  # Flag to skip the row if "Amount of fine imposed" is empty
        
        for fine_col, master_col in COLUMN_MAPPING.items():
            if master_col:  # Skip mappings with empty master columns
                fine_col_normalized = normalize_string(fine_col)
                master_col_normalized = normalize_string(master_col)
                if fine_col_normalized in fine_headers and master_col_normalized in master_headers:
                    fine_col_idx = fine_headers[fine_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    
                    # Check if "Amount of fine imposed" is empty
                    if fine_col == "Amount of fine imposed" and not master_row[master_col_idx - 1]:
                        skip_row = True
                        break  # Skip processing this row
                    fine_row_data[fine_col_idx] = master_row[master_col_idx - 1]

        # Populate the row in Fine.xlsx if not skipped
        if not skip_row and fine_row_data:
            for col_idx, value in fine_row_data.items():
                cell = sheet_fine.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align data
            row_idx += 1
            data_inserted = False

    # If no data was inserted, write "NO FINES IMPOSED FOR THE MONTH OF OCT 2024"
    if not data_inserted:
        for row in sheet_fine.iter_rows(min_row=14):
            for cell in row:
                cell.value = None
        start_column = 4  # Column D
        end_column = sheet_fine.max_column  # End of the table
        sheet_fine.merge_cells(start_row=18, start_column=start_column, 
                               end_row=18, end_column=end_column)
        message_cell = sheet_fine.cell(row=18, column=start_column)
        message_cell.value = "NO FINES IMPOSED FOR THE MONTH OF NOV 2024"
        message_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Apply borders to all cells under row 13 for columns with headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in fine_headers.values():
        for row_idx in range(14, sheet_fine.max_row + 1):
            cell = sheet_fine.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge cells only up to row 13
    for merged_range in merged_cells:
        start_row = merged_range.min_row
        if start_row <= 13:
            sheet_fine.merge_cells(str(merged_range))

    # Save the updated Fine.xlsx file
    wb_fine.save(temp_output_file)
    print(f"Data inserted, borders applied, and file saved as: {temp_output_file}")

    return temp_output_file
