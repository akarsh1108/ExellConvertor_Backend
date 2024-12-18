from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from collections import defaultdict
import os


# Column mapping
COLUMN_MAPPING = {
    "Category of workers": "",
    "Brief description of work": "Designation",
    "No. of men employed": "No. of men employed",
    "No. of women employed": "No. of women employed",
    "Rate of remuneration paid": "Full GROSS",
    "Basic wage or salary": "Basic",
    "Components of remuneration_Dearness allowance": "",
    "House Rent allowance": "HRA",
    "Other allowances": "",
    "Cash value of concessional supply of essential commodities": "",
}

def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""

def combine_headers(row_10, row_11):
    """Combine headers from two rows to handle nested headers."""
    combined_headers = []
    for col_idx, header_row_10 in enumerate(row_10):
        # Convert both header_row_10 and header_row_11 to strings if they are not None
        header_row_10 = str(header_row_10).strip() if header_row_10 else ""
        header_row_11 = str(row_11[col_idx]).strip() if col_idx < len(row_11) and row_11[col_idx] else ""
        
        if header_row_10 and header_row_11:
            combined_headers.append(f"{header_row_10}_{header_row_11}")
        elif header_row_10:
            combined_headers.append(header_row_10)
        elif header_row_11:
            combined_headers.append(header_row_11)
        else:
            combined_headers.append("")  # Empty header
    return combined_headers

def calculate_gender_counts(master_sheet):
    """Calculate the number of men and women for each designation."""
    gender_counts = defaultdict(lambda: {"Male": 0, "Female": 0})
    # Find the indices for Designation and Gender columns
    headers = [normalize_string(cell.value) for cell in master_sheet[3]]
    designation_idx = headers.index("designation")
    gender_idx = headers.index("gender")

    # Count occurrences of Male and Female by Designation
    for row in master_sheet.iter_rows(min_row=4, values_only=True):
        designation = normalize_string(row[designation_idx])
        gender = normalize_string(row[gender_idx])
        if gender in {"male", "female"}:
            gender_counts[designation][gender.capitalize()] += 1

    return gender_counts
def formD_process_excel(base_file_path, temp_master_file_path, temp_output_file):
    # Load the input workbook (accident.xlsx)
    wb_accident = load_workbook(base_file_path)
    sheet_accident = wb_accident.active  # Assuming first sheet in accident.xlsx

    # Load the master workbook (master.xlsx)
    wb_master = load_workbook(temp_master_file_path)
    sheet_master = wb_master.active  # Assuming first sheet in master.xlsx

    # Combine headers from row 10 and 11 in the accident file
    row_10 = [cell.value for cell in sheet_accident[10]]
    row_11 = [cell.value for cell in sheet_accident[11]]
    combined_headers = combine_headers(row_10, row_11)

    # Normalize combined headers
    accident_headers = {normalize_string(header): idx + 1 for idx, header in enumerate(combined_headers)}

    # Get headers from row 3 in master.xlsx and normalize
    master_headers = {normalize_string(cell.value): idx + 1 for idx, cell in enumerate(sheet_master[3])}

    # Get gender counts from the master sheet
    gender_counts = calculate_gender_counts(sheet_master)

    # Unmerge all merged cells in accident.xlsx before clearing
    merged_cells = list(sheet_accident.merged_cells)  # Store merged cell ranges
    for merged_range in merged_cells:
        sheet_accident.unmerge_cells(str(merged_range))

    # Clear content in accident.xlsx from row 13 onwards
    for row in sheet_accident.iter_rows(min_row=13):
        for cell in row:
            cell.value = None

    # Insert data from master.xlsx into accident.xlsx starting from row 13
    row_idx = 13
    total_men = 0
    total_women = 0
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):  # Data starts from row 4
        accident_row_data = {}  # Temporary storage for the accident row data
        for accident_col, master_col in COLUMN_MAPPING.items():
            if master_col:  # Skip mappings with empty master columns
                accident_col_normalized = normalize_string(accident_col)
                master_col_normalized = normalize_string(master_col)
                if accident_col_normalized in accident_headers and master_col_normalized in master_headers:
                    accident_col_idx = accident_headers[accident_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    accident_row_data[accident_col_idx] = master_row[master_col_idx - 1]

        # Populate the row in accident.xlsx
        for col_idx, value in accident_row_data.items():
            cell = sheet_accident.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align data

        # Insert "Skilled" in column A and 0 in column G
        sheet_accident.cell(row=row_idx, column=1, value="Skilled")  # Column A
        sheet_accident.cell(row=row_idx, column=7, value=0)          # Column G

        # Calculate sum of values from columns 55 to 63 (BD to BK) and store in column I
        bd_bk_sum = 0
        for col_idx in range(55, 64):  # Columns BD to BK (1-based index: 55 to 63)
            value = master_row[col_idx - 1]  # Convert to 0-based index
            bd_bk_sum += float(value) if isinstance(value, (int, float)) else 0
        
        # Write the calculated sum into column I (9th column)
        sheet_accident.cell(row=row_idx, column=9, value=bd_bk_sum)

        # Add gender counts to totals
        designation = normalize_string(sheet_accident.cell(row=row_idx, column=accident_headers.get("brief description of work")).value)
        if designation in gender_counts:
            total_men += gender_counts[designation]["Male"]
            total_women += gender_counts[designation]["Female"]

        row_idx += 1
            # Populate gender data into the accident file
    for row_idx, row in enumerate(sheet_accident.iter_rows(min_row=13, max_row=sheet_accident.max_row), start=13):
        designation_col_idx = accident_headers.get("brief description of work", None)
        men_col_idx = accident_headers.get("no. of men employed", None)
        women_col_idx = accident_headers.get("no. of women employed", None)

        if designation_col_idx and men_col_idx and women_col_idx:
            designation = normalize_string(sheet_accident.cell(row=row_idx, column=designation_col_idx).value)
            if designation in gender_counts:
                sheet_accident.cell(row=row_idx, column=men_col_idx).value = gender_counts[designation]["Male"]
                sheet_accident.cell(row=row_idx, column=women_col_idx).value = gender_counts[designation]["Female"]


    # Populate totals in D7, D8, D9
    total_workers = total_men + total_women
    sheet_accident["D7"].value = f"{total_workers}"
    sheet_accident["D8"].value = f"{total_men}"
    sheet_accident["D9"].value = f"{total_women}"

    # Apply borders to all cells under row 12 for columns with headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in accident_headers.values():
        for row_idx in range(13, sheet_accident.max_row + 1):
            cell = sheet_accident.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge cells only up to row 12
    for merged_range in merged_cells:
        start_row = merged_range.min_row
        if start_row <= 12:
            sheet_accident.merge_cells(str(merged_range))

    # Save the updated accident.xlsx file
    wb_accident.save(temp_output_file)
    print(f"Data updated with 'Skilled' in Column A, '0' in Column G, and sum in Column I. File saved to: {temp_output_file}")

    return temp_output_file
