from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

import os

# Column mapping
COLUMN_MAPPING = {
    "Sl .No": "Sr #",
    "Emp Code": "EmployeeNo",
    "Name of Workman": "Name",
    "Designation/Nature of Works Done": "Designation",
    "State": "State",
    "Location": "Location",
    "UAN":"Unique PF No.",
    "ESIC":"Esic No",
    "Unit of works done":"EMPLOYEE WORKDAYS",
    "Amount of Wages Earned_Basic+DA": "Basic",
    "Basic+DA Arrear": "Basic (Arrear)",
    "HRA": "HRA",
    "HRA Arrear": "HRA (Arrear)",
    "Special Allow.": "Special Allowance",
    "Special Allow. Arrear":"Special Allowance (Arrear)",
    "Special Allow. PF": "Special Allowance - PF",
    "Special Allow. PF Arrear": "Special Allowance - PF (Arrear)",
    "Bonus Gross": "Bonus Gross",
    "Bonus Gross Arrear": "Bonus Gross (Arrear)",
    "Incentive":"Onetime/Quarterly Incentive",
    "OT":"OT",
    "Gross pay":"GROSS",
    "Deduction_PF":"PF Employee Ceiling",
    "ESI":"ESIC Employee",
    "PT":"Professional Tax",
    "Income Tax":"Income Tax",
    "LWF":"Labour Welfare Fund Employee",
    "Salary Advance":"Salary Advance",
    "Other Deduction":"Other Deduction",
    "Total Deduction":"TOTAL DEDUCTIONS",
    "Net Amount Paid":"NET PAY",
    "Bank A/c No.":"Bank A/C",
    "Bank Name":"Bank Name",

}


def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""

def combine_headers(row_11, row_12):
    """Combine headers from two rows to handle nested headers."""
    combined_headers = []
    for col_idx, header_row_11 in enumerate(row_11):
        # Convert both header_row_11 and header_row_12 to strings if they are not None
        header_row_11 = str(header_row_11).strip() if header_row_11 else ""
        header_row_12 = str(row_12[col_idx]).strip() if col_idx < len(row_12) and row_12[col_idx] else ""
        
        if header_row_11 and header_row_12:
            combined_headers.append(f"{header_row_11}_{header_row_12}")
        elif header_row_11:
            combined_headers.append(header_row_11)
        elif header_row_12:
            combined_headers.append(header_row_12)
        else:
            combined_headers.append("")  # Empty header
    return combined_headers


def wages_process_excel(base_file_path, temp_master_file_path, temp_output_file):
    # Load the input workbook
    wb_input = load_workbook(base_file_path)
    sheet_input = wb_input.active  # Assuming the first sheet

    # Load the master workbook
    wb_master = load_workbook(temp_master_file_path)
    sheet_master = wb_master.active  # Assuming the first sheet

    # Combine and normalize headers
    row_11 = [cell.value for cell in sheet_input[11]]
    row_12 = [cell.value for cell in sheet_input[12]]
    combined_headers = combine_headers(row_11, row_12)
    input_headers = {normalize_string(header): idx + 1 for idx, header in enumerate(combined_headers)}
    master_headers = {normalize_string(cell.value): idx for idx, cell in enumerate(sheet_master[3], start=1)}

    # Unmerge all merged cells
    merged_cells = list(sheet_input.merged_cells)
    for merged_range in merged_cells:
        sheet_input.unmerge_cells(str(merged_range))

    # Clear content from row 14 onwards
    for row in sheet_input.iter_rows(min_row=14):
        for cell in row:
            cell.value = None

    # Insert data from master file
    row_idx = 14
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):
        input_row_data = {}
        for input_col, master_col in COLUMN_MAPPING.items():
            if master_col:
                input_col_normalized = normalize_string(input_col)
                master_col_normalized = normalize_string(master_col)
                if input_col_normalized in input_headers and master_col_normalized in master_headers:
                    input_col_idx = input_headers[input_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    input_row_data[input_col_idx] = master_row[master_col_idx - 1]

        for col_idx, value in input_row_data.items():
            cell = sheet_input.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Calculate and insert sums in BL and BT
        bl_sum = 0
        bt_sum = 0

        # AZ to BK (columns 52 to 63 in 1-based index)
        for col_idx in range(52, 64):
            value = master_row[col_idx - 1]  # Convert to 0-based index
            bl_sum += float(value) if isinstance(value, (int, float)) else 0

        # Bm to BP (columns 65 to 68 in 1-based index)
        for col_idx in range(65, 69):
            value = master_row[col_idx - 1]  # Convert to 0-based index
            bt_sum += float(value) if isinstance(value, (int, float)) else 0

        # Write the calculated sums
        sheet_input.cell(row=row_idx, column=23, value=bl_sum)  # W (column 23)
        sheet_input.cell(row=row_idx, column=31, value=bt_sum)  # AE (column 31)

        # Calculate and insert BU (BL - BT)
        bu_value = bl_sum - bt_sum
        sheet_input.cell(row=row_idx, column=32, value=bu_value)  # BU (column 73)

        row_idx += 1

    # Apply borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in input_headers.values():
        for row_idx in range(14, sheet_input.max_row + 1):
            cell = sheet_input.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge header cells
    for merged_range in merged_cells:
        if merged_range.min_row <= 13:
            sheet_input.merge_cells(str(merged_range))

    # Save the updated file
    wb_input.save(temp_output_file)
    print(f"Data inserted and formatted successfully. File saved as: {temp_output_file}")

    return temp_output_file
