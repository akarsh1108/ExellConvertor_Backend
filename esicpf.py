import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# Define column mapping
COLUMN_MAPPING = {
    "Sl .No": "Sr #",
    "E_Code":"EmployeeNo",
    "Name of  workman": "Name",
    "UAN No": "Unique PF No.",
    "ESIC No.": "Esic No",
    "State":"State",
    "Paid Days": "EMPLOYEE WORKDAYS",
    "Date of Leaving": "Date of Leaving",
    "Basic +DA": "Basic",
    "EPF Contribution": "PF Employee Ceiling",
    "ESIC Contribution": "ESIC Employee",
    "Location": "Location",
    "Basic + DA": "Basic",
    "Basic + DA Arrear":"Basic (Arrear)"
}

def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""
def esicpf_process_excel(esicpf_file_path, master_file_path, output_file_path):
    """Process the ESIC and PF Excel file with data from the master file."""
    # Load the input workbook (esicpf.xlsx)
    wb_esicpf = load_workbook(esicpf_file_path)
    sheet_esicpf = wb_esicpf.active  # Assuming the first sheet in esicpf.xlsx

    # Load the master workbook (master.xlsx)
    wb_master = load_workbook(master_file_path)
    sheet_master = wb_master.active  # Assuming the first sheet in master.xlsx

    # Get headers from row 11 in ESIC and PF file and normalize
    esicpf_headers = {normalize_string(cell.value): idx + 1 for idx, cell in enumerate(sheet_esicpf[10])}

    # Get headers from row 3 in master.xlsx and normalize
    master_headers = {normalize_string(cell.value): idx for idx, cell in enumerate(sheet_master[3], start=1)}

    # Unmerge all merged cells in esicpf.xlsx before clearing
    merged_cells = list(sheet_esicpf.merged_cells)  # Store merged cell ranges
    for merged_range in merged_cells:
        sheet_esicpf.unmerge_cells(str(merged_range))

    # Clear content in esicpf.xlsx from row 12 onwards
    for row in sheet_esicpf.iter_rows(min_row=11):
        for cell in row:
            cell.value = None

    # Insert data from master.xlsx into esicpf.xlsx starting from row 12
    row_idx = 11
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):  # Data starts from row 4
        esicpf_row_data = {}  # Temporary storage for the ESIC and PF row data
        for esicpf_col, master_col in COLUMN_MAPPING.items():
            if master_col:  # Skip mappings with empty master columns
                esicpf_col_normalized = normalize_string(esicpf_col)
                master_col_normalized = normalize_string(master_col)
                if esicpf_col_normalized in esicpf_headers and master_col_normalized in master_headers:
                    esicpf_col_idx = esicpf_headers[esicpf_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    esicpf_row_data[esicpf_col_idx] = master_row[master_col_idx - 1]

        # Populate the row in esicpf.xlsx
        for col_idx, value in esicpf_row_data.items():
            cell = sheet_esicpf.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align data

        row_idx += 1

    # Apply borders to all cells under row 11 for columns with headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in esicpf_headers.values():
        for row_idx in range(12, sheet_esicpf.max_row + 1):
            cell = sheet_esicpf.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge cells only up to row 11
    for merged_range in merged_cells:
        start_row = merged_range.min_row
        if start_row <= 11:
            sheet_esicpf.merge_cells(str(merged_range))

    # Save the updated esicpf.xlsx file
    wb_esicpf.save(output_file_path)
    print(f"Data inserted and borders applied successfully. New file saved as: {output_file_path}")

    return output_file_path
