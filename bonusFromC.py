import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# Define column mapping
COLUMN_MAPPING = {
    "Sl. No.": "Sr #",
    "Emp Code": "EmployeeNo",
    "Name of the employee": "Name",
    "Father's name": "Father Name",
    "Whether he has completed 15 years of age at the beginning of the accounting month": "",
    "Designation ": "Designation",
    "No. of days worked in the month": "EMPLOYEE WORKDAYS",
    "Total salary or wage in respect of the accounting Month": "Basic",
    "Amount of bonus payable under section 10 or section 11, as the case may be": "Bonus Gross",
    "Puja bonus or other customary bonus paid during the accounting Month": "",
    "Interim bonus or bonus paid in advance":  "Bonus Gross",
    "Amount of Income-tax deduced":"Income Tax",
    "Deduction on account of financial loss, if any, caused by misconduct of the employee":"",
    "Total sum deducted under Columns 9, 10, 10A and 11":"Bonus Gross",
    "Net amount payable (Column 8 minus Column 12)":"Bonus Gross",
    "Amount actually paid":"Bonus Gross",
    "Date on which paid":"Salary Processed Month",
    "Signature/Thumb impression of the employee":"",
    "State":"State",
    "Location":"Location",
    "Basic":"Basic",
    "Basic (Arrear)":"Basic (Arrear)",
    "Bonus Gross":"Bonus Gross",
}

def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""

def combine_headers(row_11, row_12):
    """Combine headers from two rows to handle nested headers."""
    combined_headers = []
    for col_idx, header_row_11 in enumerate(row_11):
        header_row_12 = row_12[col_idx] if col_idx < len(row_12) else None
        if header_row_11 and header_row_12:
            combined_headers.append(f"{header_row_11.strip()}_{header_row_12.strip()}")
        elif header_row_11:
            combined_headers.append(header_row_11.strip())
        elif header_row_12:
            combined_headers.append(header_row_12.strip())
        else:
            combined_headers.append("")  # Empty header
    return combined_headers


def bonus_process_excel(bonus_file_path, master_file_path, output_file_path):
    """Process the Bonus Form C Excel file with data from the master file."""
    # Load the input workbook (BonusFormC.xlsx)
    wb_bonus = load_workbook(bonus_file_path)
    sheet_bonus = wb_bonus.active  # Assuming the first sheet in BonusFormC.xlsx

    # Load the master workbook (master.xlsx)
    wb_master = load_workbook(master_file_path)
    sheet_master = wb_master.active  # Assuming the first sheet in master.xlsx

    # Combine headers from row 11 and 12 in the Bonus Form C file
    row_11 = [cell.value for cell in sheet_bonus[11]]
    row_12 = [cell.value for cell in sheet_bonus[12]]
    combined_headers = combine_headers(row_11, row_12)

    # Normalize combined headers
    bonus_headers = {normalize_string(header): idx + 1 for idx, header in enumerate(combined_headers)}

    # Get headers from row 3 in master.xlsx and normalize
    master_headers = {normalize_string(cell.value): idx for idx, cell in enumerate(sheet_master[3], start=1)}

    # Unmerge all merged cells in BonusFormC.xlsx before clearing
    merged_cells = list(sheet_bonus.merged_cells)  # Store merged cell ranges
    for merged_range in merged_cells:
        sheet_bonus.unmerge_cells(str(merged_range))

    # Clear content in BonusFormC.xlsx from row 14 onwards
    for row in sheet_bonus.iter_rows(min_row=14):
        for cell in row:
            cell.value = None

    # Insert data from master.xlsx into BonusFormC.xlsx starting from row 14
    row_idx = 14
    column_r_index = 18  # Column R corresponds to the 18th column
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):  # Data starts from row 4
        bonus_row_data = {}  # Temporary storage for the bonus row data
        for bonus_col, master_col in COLUMN_MAPPING.items():
            if master_col:  # Skip mappings with empty master columns
                bonus_col_normalized = normalize_string(bonus_col)
                master_col_normalized = normalize_string(master_col)
                if bonus_col_normalized in bonus_headers and master_col_normalized in master_headers:
                    bonus_col_idx = bonus_headers[bonus_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    bonus_row_data[bonus_col_idx] = master_row[master_col_idx - 1]

        # Populate the row in BonusFormC.xlsx
        for col_idx, value in bonus_row_data.items():
            cell = sheet_bonus.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align data

        # Write "Paid on every 7th day of each month" in Column R
        sheet_bonus.cell(row=row_idx, column=column_r_index, value="Paid on every 7th day of each month")

        row_idx += 1

    # Apply borders to all cells under row 13 for columns with headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in bonus_headers.values():
        for row_idx in range(14, sheet_bonus.max_row + 1):
            cell = sheet_bonus.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge cells only up to row 13
    for merged_range in merged_cells:
        start_row = merged_range.min_row
        if start_row <= 13:
            sheet_bonus.merge_cells(str(merged_range))

    # Save the updated BonusFormC.xlsx file
    wb_bonus.save(output_file_path)
    print(f"Data inserted, 'Paid on every 7th day of each month' added in column R. File saved as: {output_file_path}")

    return output_file_path
