import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# Define column mapping
COLUMN_MAPPING = {
    "Sl.No": "Sr #",
    "Name Of the Employee": "Name",
    "Father's/Husband's  Name": "Father Name",
    "Sex": "Gender",
    "Dasignation / nature of employment": "Designation",
    "Dates on which overtime worked": "",
    "Normal rate of wages": "Basic",
    "overtime rate of wages": "OT",
    "overtime earnings": "",
    "Date on which overtime wages paid": "",
    "Remarks": "",
}

def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""

def combine_headers(row_11, row_12):
    """Combine headers from two rows to handle nested headers."""
    combined_headers = []
    for col_idx, header_row_11 in enumerate(row_11):
        # Convert both header_row_11 and header_row_12 to strings if they are not None
        header_row_11 = str(header_row_11).strip() if header_row_11 else ""
        header_row_12 = str(row_12[col_idx]).strip() if col_idx < len(row_12) and row_12[col_idx] else ""
        
        if header_row_11 and header_row_12:
            combined_headers.append(f"{header_row_11}_{header_row_12}")
        elif header_row_11:
            combined_headers.append(header_row_11)
        elif header_row_12:
            combined_headers.append(header_row_12)
        else:
            combined_headers.append("")  # Empty header
    return combined_headers


def overtime_process_excel(base_file_path, temp_master_file_path, temp_output_file):
    # Load the input workbook
    wb_accident = load_workbook(base_file_path)
    sheet_accident = wb_accident.active  # Assuming the first sheet

    # Load the master workbook
    wb_master = load_workbook(temp_master_file_path)
    sheet_master = wb_master.active  # Assuming the first sheet

    # Combine headers from row 11 and 12 in the input file
    row_11 = [cell.value for cell in sheet_accident[11]]
    row_12 = [cell.value for cell in sheet_accident[12]]
    combined_headers = combine_headers(row_11, row_12)

    # Normalize combined headers
    accident_headers = {normalize_string(header): idx + 1 for idx, header in enumerate(combined_headers)}

    # Get headers from row 3 in master.xlsx and normalize
    master_headers = {normalize_string(cell.value): idx for idx, cell in enumerate(sheet_master[3], start=1)}

    # Unmerge all merged cells in the input file before clearing
    merged_cells = list(sheet_accident.merged_cells)  # Store merged cell ranges
    for merged_range in merged_cells:
        sheet_accident.unmerge_cells(str(merged_range))

    # Clear content in the input file from row 14 onwards
    for row in sheet_accident.iter_rows(min_row=14):
        for cell in row:
            cell.value = None

    # Insert data from master.xlsx into the input file starting from row 14
    row_idx = 14
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):  # Data starts from row 4
        accident_row_data = {}  # Temporary storage for the row data
        for accident_col, master_col in COLUMN_MAPPING.items():
            if master_col:  # Skip mappings with empty master columns
                accident_col_normalized = normalize_string(accident_col)
                master_col_normalized = normalize_string(master_col)
                if accident_col_normalized in accident_headers and master_col_normalized in master_headers:
                    accident_col_idx = accident_headers[accident_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    accident_row_data[accident_col_idx] = master_row[master_col_idx - 1]

        # Populate the row in the input file
        for col_idx, value in accident_row_data.items():
            cell = sheet_accident.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align data
        
        row_idx += 1

    # Apply borders to all cells under row 12 for columns with headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in accident_headers.values():
        for row_idx in range(13, sheet_accident.max_row + 1):
            cell = sheet_accident.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge cells only up to row 12
    for merged_range in merged_cells:
        start_row = merged_range.min_row
        if start_row <= 13:
            sheet_accident.merge_cells(str(merged_range))

    # Save the updated file
    wb_accident.save(temp_output_file)
    print(f"Data inserted and borders applied successfully. File saved as: {temp_output_file}")

    return temp_output_file

