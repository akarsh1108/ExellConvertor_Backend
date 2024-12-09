from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
import os

# File paths
MASTER_FILE_PATH = os.path.join("input", "master.xlsx")
STATIC_FILE_PATH = os.path.join("input", "damage.xlsx")
OUTPUT_FILE_PATH = os.path.join("output", "damage.xlsx")

# Column mapping
COLUMN_MAPPING = {
    "Sl .No": "Sr #",
    "Name Of Employee": "Name",
    "Father's/ Husband's  Name": "Father Name",
    "Nature of Employment/ Designation": "Designation",
    "Particulars of Damages or Loss": "",
    "Date of Damage or Loss": "",
    "Whether workman showed cause against deduction": "",
    "Name of person in whose presence employee's explanation was heard": "",
    "Amount  of deduction imposed": "",
    "Date of Recovery_No. of Instalments": "",
    "First Instalments": "",
    "Last Instalments": "",
    "Remarks": "",
}


def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""

def combine_headers(row_11, row_12):
    """Combine headers from two rows to handle nested headers."""
    combined_headers = []
    for col_idx, header_row_11 in enumerate(row_11):
        # Convert both header_row_11 and header_row_12 to strings if they are not None
        header_row_11 = str(header_row_11).strip() if header_row_11 else ""
        header_row_12 = str(row_12[col_idx]).strip() if col_idx < len(row_12) and row_12[col_idx] else ""
        
        if header_row_11 and header_row_12:
            combined_headers.append(f"{header_row_11}_{header_row_12}")
        elif header_row_11:
            combined_headers.append(header_row_11)
        elif header_row_12:
            combined_headers.append(header_row_12)
        else:
            combined_headers.append("")  # Empty header
    return combined_headers


def damage_process_excel(damage_file_path, master_file_path, output_file_path):
    """Process the Damage Excel file with data from the master file."""
    # Load the input workbook (damage.xlsx)
    wb_damage = load_workbook(damage_file_path)
    sheet_damage = wb_damage.active  # Assuming the first sheet in damage.xlsx

    # Load the master workbook (master.xlsx)
    wb_master = load_workbook(master_file_path)
    sheet_master = wb_master.active  # Assuming the first sheet in master.xlsx

    # Combine headers from row 11 and 12 in the damage file
    row_11 = [cell.value for cell in sheet_damage[11]]
    row_12 = [cell.value for cell in sheet_damage[12]]
    combined_headers = combine_headers(row_11, row_12)

    # Normalize combined headers
    damage_headers = {normalize_string(header): idx + 1 for idx, header in enumerate(combined_headers)}

    # Get headers from row 3 in master.xlsx and normalize
    master_headers = {normalize_string(cell.value): idx for idx, cell in enumerate(sheet_master[3], start=1)}

    # Unmerge all merged cells in damage.xlsx before clearing
    merged_cells = list(sheet_damage.merged_cells)  # Store merged cell ranges
    for merged_range in merged_cells:
        sheet_damage.unmerge_cells(str(merged_range))

    # Clear content in damage.xlsx from row 14 onwards
    for row in sheet_damage.iter_rows(min_row=14):
        for cell in row:
            cell.value = None

    # Insert data from master.xlsx into damage.xlsx starting from row 14
    row_idx = 14
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):  # Data starts from row 4
        damage_row_data = {}  # Temporary storage for the damage row data
        for damage_col, master_col in COLUMN_MAPPING.items():
            if master_col:  # Skip mappings with empty master columns
                damage_col_normalized = normalize_string(damage_col)
                master_col_normalized = normalize_string(master_col)
                if damage_col_normalized in damage_headers and master_col_normalized in master_headers:
                    damage_col_idx = damage_headers[damage_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    damage_row_data[damage_col_idx] = master_row[master_col_idx - 1]

        # Populate the row in damage.xlsx
        for col_idx, value in damage_row_data.items():
            cell = sheet_damage.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align data
        
        row_idx += 1

    # Apply borders to all cells under row 12 for columns with headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in damage_headers.values():
        for row_idx in range(14, sheet_damage.max_row + 1):
            cell = sheet_damage.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge cells only up to row 13
    for merged_range in merged_cells:
        start_row = merged_range.min_row
        if start_row <= 13:
            sheet_damage.merge_cells(str(merged_range))

    # Save the updated damage.xlsx file
    wb_damage.save(output_file_path)
    print(f"Data inserted and borders applied successfully. New file saved as: {output_file_path}")

    return output_file_path
