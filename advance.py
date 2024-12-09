import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# Define column mapping
COLUMN_MAPPING = {
    "Sl.No": "Sr #",
    "Name Of the Employee": "Name",
    "Father's/ Husband's  Name": "Father Name",
    "Nature of Employment/ Designation": "Designation",
    "Wage Peroid and wages Payable": "",
    "Date and amount of advance given": "",
    "Purpose(s) for Which advance mace": "",
    "No of Instalments by which advance to be repaid": "",
    "Date and amount of each instalment was paid": "",
    "Date on which last instalment was repaid": "",
    "Remarks": "",
}


def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""

def advance_process_excel(advance_file_path, master_file_path, output_file_path):
    # Load the input workbook (Advance.xlsx)
    wb_advance = load_workbook(advance_file_path)
    sheet_advance = wb_advance.active  # Assuming first sheet in Advance.xlsx

    # Load the master workbook (master.xlsx)
    wb_master = load_workbook(master_file_path)
    sheet_master = wb_master.active  # Assuming first sheet in master.xlsx

    # Get headers from row 11 in Advance.xlsx and normalize
    advance_headers = {normalize_string(cell.value): idx + 1 for idx, cell in enumerate(sheet_advance[11])}

    # Get headers from row 3 in master.xlsx and normalize
    master_headers = {normalize_string(cell.value): idx for idx, cell in enumerate(sheet_master[3], start=1)}

    # Unmerge all merged cells in Advance.xlsx before clearing
    merged_cells = list(sheet_advance.merged_cells)
    for merged_range in merged_cells:
        sheet_advance.unmerge_cells(str(merged_range))

    # Clear content in Advance.xlsx from row 14 onwards
    for row in sheet_advance.iter_rows(min_row=14):
        for cell in row:
            cell.value = None

    # Insert data from master.xlsx into Advance.xlsx starting from row 14
    row_idx = 14
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):  # Data starts from row 4
        advance_row_data = {}  # Temporary storage for the advance row data
        for advance_col, master_col in COLUMN_MAPPING.items():
            if master_col:  # Skip mappings with empty master columns
                advance_col_normalized = normalize_string(advance_col)
                master_col_normalized = normalize_string(master_col)
                if advance_col_normalized in advance_headers and master_col_normalized in master_headers:
                    advance_col_idx = advance_headers[advance_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    advance_row_data[advance_col_idx] = master_row[master_col_idx - 1]

        # Populate the row in Advance.xlsx
        for col_idx, value in advance_row_data.items():
            cell = sheet_advance.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_idx += 1

    # Apply borders to all cells under row 13 for columns with headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in advance_headers.values():
        for row_idx in range(14, sheet_advance.max_row + 1):
            cell = sheet_advance.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge cells only up to row 13
    for merged_range in merged_cells:
        start_row = merged_range.min_row
        if start_row <= 13:
            sheet_advance.merge_cells(str(merged_range))

    # Remove the rightmost column
    rightmost_column = max(advance_headers.values())
    sheet_advance.delete_cols(rightmost_column)

    # Remove the two bottommost rows
    bottommost_row = sheet_advance.max_row
    sheet_advance.delete_rows(bottommost_row, 2)

    # Save the updated Advance.xlsx file
    wb_advance.save(output_file_path)
    print(f"Data inserted, rightmost column removed, and bottommost rows deleted. New file saved as: {output_file_path}")

    return output_file_path
