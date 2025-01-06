# import csv
# import os
# from openpyxl import load_workbook
# import fitz  # PyMuPDF


# def highlight_identifiers_in_pdf(input_pdf_path, output_pdf_path, identifiers):
#     """
#     Highlight specific identifiers on all pages of a PDF.

#     Args:
#         input_pdf_path (str): Path to the input PDF file.
#         output_pdf_path (str): Path to save the highlighted PDF.
#         identifiers (list): List of identifiers to highlight.

#     Returns:
#         str: Path to the highlighted PDF.
#     """
#     # Open the PDF
#     doc = fitz.open(input_pdf_path)

#     # Iterate over all pages
#     for page_index in range(len(doc)):
#         # Select the current page
#         page = doc[page_index]

#         # Highlight each identifier
#         for identifier in identifiers:
#             text_instances = page.search_for(identifier)
#             # print(f"Page {page_index + 1}: Found instances of '{identifier}': {text_instances}")
#             for inst in text_instances:
#                 # Adjust the bounding box for the found text
#                 inst.y0 -= 715
#                 inst.y1 += 35 
#                 inst.x0 -= 5
#                 inst.x1 += 5

#                 # Add a highlight annotation
#                 rect = fitz.Rect(inst)
#                 highlight = page.add_rect_annot(rect)
#                 highlight.set_colors(stroke=(1, 1, 0), fill=(1, 1, 0))  # Yellow color
#                 highlight.set_opacity(0.3)  # Set transparency to 30%
#                 highlight.update()

#     # Save the modified PDF
#     doc.save(output_pdf_path, garbage=4, deflate=True)

#     return output_pdf_path
# # Example usage
# if __name__ == "__main__":
#     # Input and output file paths
#     input_pdf = "ECR.pdf"
#     output_pdf = "ECR_highlighted_result.pdf"

#    # Path to your Excel file
#     xlsx_file_path = os.path.join("input","master.xlsx")

#     # List to store the extracted values
#     identifiers_to_highlight = []

#     # Load the workbook and select the active sheet
#     workbook = load_workbook(filename=xlsx_file_path)
#     sheet = workbook.active

#     # Locate the column index for the header "Unique PF No."
#     header_row = 3
#     unique_pf_column_index = None

#     for col in sheet.iter_cols(min_row=header_row, max_row=header_row):
#         if col[0].value == "Unique PF No.":
#             unique_pf_column_index = col[0].column
#             break

#     if unique_pf_column_index is None:
#         print("Header 'Unique PF No.' not found in the file.")
#     else:
#         # Extract all the values from the specified column starting from row 4
#         for row in sheet.iter_rows(min_row=4, min_col=unique_pf_column_index, max_col=unique_pf_column_index):
#             cell_value = row[0].value
#             if cell_value is not None:  # Add only non-empty values
#                 identifiers_to_highlight.append(cell_value)

        
#     # Page index to process (0-based, e.g., 1 for the second page)
#     page_to_process = 1

#     # Highlight identifiers
#     result_pdf = highlight_identifiers_in_pdf(
#         input_pdf_path=input_pdf,
#         output_pdf_path=output_pdf,
#         identifiers=identifiers_to_highlight,
        
#     )

#     # print(f"Highlighted PDF saved at: {result_pdf}")


import fitz  # PyMuPDF


def highlight_identifiers_in_pdf(input_pdf_path, output_pdf_path, identifiers):
    """
    Highlight specific identifiers on all pages of a PDF.

    Args:
        input_pdf_path (str): Path to the input PDF file.
        output_pdf_path (str): Path to save the highlighted PDF.
        identifiers (list): List of identifiers to highlight.

    Returns:
        str: Path to the highlighted PDF.
    """
    doc = fitz.open(input_pdf_path)
    for page_index in range(len(doc)):
        page = doc[page_index]
        for identifier in identifiers:
            text_instances = page.search_for(identifier)
            for inst in text_instances:
                inst.y0 -= 715
                inst.y1 += 35
                inst.x0 -= 5
                inst.x1 += 5
                rect = fitz.Rect(inst)
                highlight = page.add_rect_annot(rect)
                highlight.set_colors(stroke=(1, 1, 0), fill=(1, 1, 0))  # Yellow color
                highlight.set_opacity(0.3)
                highlight.update()
    doc.save(output_pdf_path, garbage=4, deflate=True)
    return output_pdf_path

from openpyxl import load_workbook


def extract_identifiers_from_excel(file_path, header="Unique PF No.", header_row=3, start_row=4):
    """
    Extract identifiers from a specific column in an Excel file.

    Args:
        file_path (str): Path to the Excel file.
        header (str): Header name to locate the column.
        header_row (int): Row number where the header is located.
        start_row (int): Row number to start extracting data.

    Returns:
        list: List of extracted identifiers.
    """
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    identifiers = []

    # Locate the column index for the specified header
    unique_pf_column_index = None
    for col in sheet.iter_cols(min_row=header_row, max_row=header_row):
        if col[0].value == header:
            unique_pf_column_index = col[0].column
            break

    if unique_pf_column_index is None:
        raise ValueError(f"Header '{header}' not found in the file.")
    
    # Extract all the values from the specified column starting from the specified row
    for row in sheet.iter_rows(min_row=start_row, min_col=unique_pf_column_index, max_col=unique_pf_column_index):
        cell_value = row[0].value
        if cell_value is not None:  # Add only non-empty values
            identifiers.append(cell_value)

    return identifiers

