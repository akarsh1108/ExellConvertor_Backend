import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from pyxlsb import open_workbook as open_xlsb

# File paths
MASTER_FILE_PATH = os.path.join("input", "master.xlsb")
STATIC_FILE_PATH = os.path.join("input", "Workmen.xlsx")
OUTPUT_FILE_PATH = os.path.join("output", "workmen_updated.xlsx")

# Column mapping
COLUMN_MAPPING = {
    "Sl.No.": "Sr #",
    "Emp Code": "EmployeeNo",
    "Name and Surname of Workmen": "Name",
    "Age and Sex": "Gender",
    "Father's/ Husband's Name": "Father Name",
    "Nature of Employment/ Designation": "Designation",
    "Permanent Home Address of Workmen (Village and Tahsil/ Taluk and District)": "",
    "Local Address": "",
    "Date of Commencement of  Employment": "Date of Joining",
    "Signature or Thumb- Impression of Workmen": "",
    "Date of Termination of Employment": "",
}

# Utility functions
def normalize_string(value):
    """Normalize string by stripping spaces and converting to lowercase."""
    return value.strip().lower() if value else ""

def combine_headers(row_1, row_2):
    """Combine headers from two rows to handle nested headers."""
    combined_headers = []
    for col_idx, header_1 in enumerate(row_1):
        header_2 = row_2[col_idx] if col_idx < len(row_2) else None
        if header_1 and header_2:
            combined_headers.append(f"{header_1.strip()}_{header_2.strip()}")
        elif header_1:
            combined_headers.append(header_1.strip())
        elif header_2:
            combined_headers.append(header_2.strip())
        else:
            combined_headers.append("")  # Empty header
    return combined_headers

def read_xlsb_with_headers(file_path, header_row_index):
    """Reads an .xlsb file and extracts headers and data starting from a specified row."""
    data = []
    headers = []
    with open_xlsb(file_path) as wb:
        sheet = wb.get_sheet(1)  # Get the first sheet
        for row_index, row in enumerate(sheet.rows()):
            row_values = [cell.v for cell in row]
            if row_index + 1 == header_row_index:
                headers = row_values
            elif row_index + 1 > header_row_index:
                data.append(dict(zip(headers, row_values)))
    return headers, data

def read_excel_with_headers(file_path, header_row_index):
    """Reads an Excel file (.xlsx or .xlsb) and extracts headers and data."""
    if file_path.endswith(".xlsb"):
        return read_xlsb_with_headers(file_path, header_row_index)

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[header_row_index]]
    data = [
        {headers[col]: cell for col, cell in enumerate(row)}
        for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True)
    ]
    return headers, data

# Main processing function
def workmen_process_excel(base_file_path, temp_master_file_path, temp_output_file):
    # Load the input workbook
    wb_accident = load_workbook(base_file_path)
    sheet_accident = wb_accident.active  # Assuming the first sheet

    # Load the master workbook
    wb_master = load_workbook(temp_master_file_path)
    sheet_master = wb_master.active  # Assuming the first sheet

    # Combine headers from row 11 and 12 in the input file
    row_11 = [cell.value for cell in sheet_accident[10]]
    combined_headers = row_11

    # Normalize combined headers
    accident_headers = {normalize_string(header): idx + 1 for idx, header in enumerate(combined_headers)}

    # Get headers from row 3 in master.xlsx and normalize
    master_headers = {normalize_string(cell.value): idx for idx, cell in enumerate(sheet_master[3], start=1)}

    # Unmerge all merged cells in the input file before clearing
    merged_cells = list(sheet_accident.merged_cells)  # Store merged cell ranges
    for merged_range in merged_cells:
        sheet_accident.unmerge_cells(str(merged_range))

    # Clear content in the input file from row 14 onwards
    for row in sheet_accident.iter_rows(min_row=11):
        for cell in row:
            cell.value = None

    # Insert data from master.xlsx into the input file starting from row 14
    row_idx = 11
    for master_row in sheet_master.iter_rows(min_row=4, values_only=True):  # Data starts from row 4
        accident_row_data = {}  # Temporary storage for the row data
        for accident_col, master_col in COLUMN_MAPPING.items():
            if master_col:  # Skip mappings with empty master columns
                accident_col_normalized = normalize_string(accident_col)
                master_col_normalized = normalize_string(master_col)
                if accident_col_normalized in accident_headers and master_col_normalized in master_headers:
                    accident_col_idx = accident_headers[accident_col_normalized]
                    master_col_idx = master_headers[master_col_normalized]
                    accident_row_data[accident_col_idx] = master_row[master_col_idx - 1]

        # Populate the row in the input file
        for col_idx, value in accident_row_data.items():
            cell = sheet_accident.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align data
        
        row_idx += 1

    # Write "Transferred" in column K for rows 12, 13, 15, and 16
    column_k_index = 11  # K is the 11th column
    rows_to_update = [12, 13, 15, 16]

    for row in rows_to_update:
        cell = sheet_accident.cell(row=row, column=column_k_index)
        cell.value = "Transferred"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=False)  # Optional: Make the text bold for emphasis

    # Apply borders to all cells under row 12 for columns with headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col_idx in accident_headers.values():
        for row_idx in range(13, sheet_accident.max_row + 1):
            cell = sheet_accident.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    # Re-merge cells only up to row 12
    for merged_range in merged_cells:
        start_row = merged_range.min_row
        if start_row <= 10:
            sheet_accident.merge_cells(str(merged_range))

    # **Trim the rightmost column**
    max_col = sheet_accident.max_column
    sheet_accident.delete_cols(max_col)

    # Save the updated file
    wb_accident.save(temp_output_file)
    print(f"Data inserted, 'Transferred' written in column K, borders applied, and rightmost column trimmed successfully. File saved as: {temp_output_file}")

    return temp_output_file
